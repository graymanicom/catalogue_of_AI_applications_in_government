"""
Format in the same way as the original excel document
"""

# ******************* Arguments ********************************
import argparse
import os
parser = argparse.ArgumentParser(description="Filter dataset based on specified criteria and save to Excel and CSV formats.")
parser.add_argument('input_to_format', type=str, help="Path to Dataset_sorted_step_3.xlsx or Shortlist.xlsx", default = os.getcwd() + '/Dataset_sorted_step_3.xlsx')
parser.add_argument('input_from_format', type=str, help="Path to directory containing PSTW_Dataset.xlsx", default = os.getcwd() + '/PSTW_Dataset.xlsx')
parser.add_argument('output_directory', type=str, help="Directory to save the output files.",  default=os.getcwd())
args = parser.parse_args()

# ******************* Parameters ********************************
output_dataset_name = 'Curated_dataset_shortlist.xlsx'

# ******************* Packages ********************************
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Border, Alignment
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from copy import copy

# ******************* Packages ********************************
# Function to copy cell formatting
def copy_cell_format(source_cell, target_cell):
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = copy(source_cell.number_format)
        target_cell.protection = copy(source_cell.protection)
        target_cell.alignment = copy(source_cell.alignment)
    if source_cell.hyperlink:
        target_cell.hyperlink = source_cell.hyperlink
    if source_cell.comment:
        target_cell.comment = source_cell.comment

# ******************* Code ********************************
# Load the source workbook and worksheet
if 'Dataset_sorted_step_3.xlsx' in args.input_to_format:
    to_format = args.input_to_format
elif 'shortlist.xlsx' in args.input_to_format:
    to_format = args.input_to_format
else:
    print("Using Dataset_sorted_step_3 not shortlist")
    to_format = args.input_to_format + '/Dataset_sorted_step_3.xlsx'

if 'PSTW_Dataset.xlsx' in args.input_from_format:
    from_format = args.input_from_format
else:
    from_format = args.input_from_format + '/PSTW_Dataset.xlsx'

source_wb = openpyxl.load_workbook(from_format)
source_ws = source_wb.active

# Load the data to be written to the new Excel file
df_to_format = pd.read_excel(to_format)

# Create a new workbook and select the active worksheet
new_wb = openpyxl.Workbook()
new_ws = new_wb.active

# Write the filtered data to the new worksheet
for r_idx, row in enumerate(dataframe_to_rows(df_to_format, index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        new_ws.cell(row=r_idx, column=c_idx, value=value)

# Copy the formatting from the source worksheet to the new worksheet
for col in range(1, 47):  # Assuming you want to copy formatting for the first 46 columns
    col_letter = get_column_letter(col)
    for row in range(1, source_ws.max_row + 1):
        source_cell = source_ws[f'{col_letter}{row}']
        new_cell = new_ws[f'{col_letter}{row}']
        copy_cell_format(source_cell, new_cell)

# Copy conditional formatting
for cf_range in source_ws.conditional_formatting._cf_rules:
    for rule in source_ws.conditional_formatting[cf_range]:
        new_ws.conditional_formatting.add(cf_range, rule)

# Save the new workbook
new_excel_path = args.output_directory + '/' + output_dataset_name
new_wb.save(new_excel_path)

print(f"Curated data with formatting saved to: {new_excel_path}")
